import tkinter as tk
from tkinter import filedialog, messagebox
import openpyxl
from collections import defaultdict
import os
from openpyxl.styles import Border, Side, Alignment, Font, PatternFill

# 파일 경로 전역 변수
input_file = ""
output_file = ""

def create_textbook_report(ws, name, textbooks, detail_text):
    # 스타일 설정
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                        top=Side(style='thin'), bottom=Side(style='thin'))
    header_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    
    # 제목 행
    ws.merge_cells('A1:C1')
    ws['A1'] = "교재비 내역서"
    ws['A1'].font = Font(bold=True)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws['A1'].fill = header_fill
    for col in range(1, 4):
        ws.cell(row=1, column=col).border = thin_border

    # 이름 행
    ws.merge_cells('A2:B2')
    ws['A2'] = "이름"
    ws['C2'] = name
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws['C2'].alignment = Alignment(horizontal='center', vertical='center')
    for col in range(1, 4):
        ws.cell(row=2, column=col).border = thin_border
        ws.cell(row=2, column=col).fill = gray_fill

    # 헤더 행
    headers = ["교재명", "정가", "할인금액"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col)
        cell.value = header
        cell.border = thin_border
        cell.fill = gray_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
    
    # 데이터 입력
    row = 4
    total_amount = 0
    book_count = 0
    
    for textbook, info in textbooks.items():
        cell = ws.cell(row=row, column=1)
        cell.value = textbook
        cell.border = thin_border
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        
        if info['original']:
            price_cell = ws.cell(row=row, column=2, value=f"₩{info['original']:,}")
            price_cell.font = Font(strike=True)
        else:
            price_cell = ws.cell(row=row, column=2, value="-")
        price_cell.border = thin_border
        price_cell.alignment = Alignment(horizontal='right', vertical='center')
        
        amount_cell = ws.cell(row=row, column=3, value=f"₩{info['discounted']:,}")
        amount_cell.border = thin_border
        amount_cell.alignment = Alignment(horizontal='right', vertical='center')
        
        total_amount += info['discounted']
        book_count += 1
        row += 1
    
    # 빈 행 채우기
    while row < 12:
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(vertical='center')
        row += 1
    
    # 합계 행
    sum_row = row
    ws.merge_cells(f'A{sum_row}:B{sum_row}')
    sum_text = ws.cell(row=sum_row, column=1)
    sum_text.value = f"\"{book_count}\" 권 / 계(월):"
    sum_text.alignment = Alignment(horizontal='right', vertical='center')
    sum_text.border = thin_border
    
    total_cell = ws.cell(row=sum_row, column=3)
    total_cell.value = f"₩{total_amount:,}"
    total_cell.border = thin_border
    total_cell.alignment = Alignment(horizontal='right', vertical='center')
    
    # 상세 내용
    detail_row = sum_row + 1
    ws.merge_cells(f'A{detail_row}:C{detail_row+1}')
    detail_cell = ws.cell(row=detail_row, column=1)
    detail_cell.value = f"\"{detail_text}\""
    detail_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    detail_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    # 병합된 셀의 모든 셀에 테두리 적용
    for row in range(detail_row, detail_row + 2):
        for col in range(1, 4):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
    
    # 열 너비 및 행 높이 조정
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15
    ws.column_dimensions['C'].width = 15
    
    for i in range(1, detail_row + 2):
        ws.row_dimensions[i].height = 25

def process_files():
    if not input_file or not output_file:
        messagebox.showerror("오류", "입력 파일과 출력 파일을 선택해주세요.")
        return
        
    try:
        wb_input = openpyxl.load_workbook(input_file)
        ws_input = wb_input.active
        wb_output = openpyxl.Workbook()
        
        data = defaultdict(lambda: {"textbooks": {}})
        
        start_row = int(start_row_entry.get())
        name_col = int(name_col_entry.get()) - 1
        amount_col = int(amount_col_entry.get()) - 1
        textbook_col = int(textbook_col_entry.get()) - 1
        
        for row in ws_input.iter_rows(min_row=start_row):
            name = row[name_col].value
            if not name:
                continue
                
            textbook = row[textbook_col].value
            amount = row[amount_col].value
            
            if "(총)" in str(textbook):
                original_price = int(amount * 10/9)
                discounted_price = amount
            else:
                original_price = None
                discounted_price = amount
                
            if textbook and amount:
                data[name]["textbooks"][textbook] = {
                    "original": original_price,
                    "discounted": discounted_price
                }
        
        for name, info in data.items():
            ws = wb_output.create_sheet(title=name)
            create_textbook_report(ws, name, info["textbooks"], 
                                 detail_textbox.get("1.0", tk.END).strip())
        
        if 'Sheet' in wb_output.sheetnames:
            wb_output.remove(wb_output['Sheet'])
        
        wb_output.save(output_file)
        messagebox.showinfo("완료", "교재비 내역서가 생성되었습니다.")
        
    except Exception as e:
        messagebox.showerror("오류", f"처리 중 오류가 발생했습니다: {str(e)}")

def select_input_file():
    global input_file
    input_file = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    input_file_label.config(text=f": {input_file}")

def select_output_file():
    global output_file
    output_file = filedialog.asksaveasfilename(defaultextension=".xlsx",
                                              filetypes=[("Excel files", "*.xlsx")])
    output_file_label.config(text=f": {output_file}")

# GUI 생성
root = tk.Tk()
root.title("BookConv_Ho")

main_frame = tk.Frame(root, padx=10, pady=10)
main_frame.pack(expand=True, fill='both')

col_frame = tk.Frame(main_frame)
col_frame.pack(fill='x', pady=(0, 10))

entries = [
    ("이름 열 번호:", "2"),
    ("금액 열 번호:", "11"),
    ("교재명 열 번호:", "8"),
    ("시작 행 번호:", "5")
]

entry_widgets = {}
for i, (label_text, default_value) in enumerate(entries):
    label = tk.Label(col_frame, text=label_text)
    label.grid(row=i, column=0, padx=5, pady=2, sticky='e')
    entry = tk.Entry(col_frame, width=5)
    entry.insert(0, default_value)
    entry.grid(row=i, column=1, padx=5, pady=2, sticky='w')
    entry_widgets[label_text] = entry

name_col_entry = entry_widgets["이름 열 번호:"]
amount_col_entry = entry_widgets["금액 열 번호:"]
textbook_col_entry = entry_widgets["교재명 열 번호:"]
start_row_entry = entry_widgets["시작 행 번호:"]

detail_frame = tk.Frame(main_frame)
detail_frame.pack(fill='x', pady=10)
tk.Label(detail_frame, text="상세 내용:").pack(anchor='w')
detail_textbox = tk.Text(detail_frame, height=4, width=50)
detail_textbox.insert("1.0", "입금금액은 지급되는대로 현금결제 / 계좌이체 해주시기 바랍니다 ★ ☞ 계좌번호 : 신한 140 - 013 - 667425 (주)")
detail_textbox.pack(fill='x')

file_frame = tk.Frame(main_frame)
file_frame.pack(fill='x', pady=10)

input_file_button = tk.Button(file_frame, text="입력 파일 선택", command=select_input_file)
input_file_button.grid(row=0, column=0, padx=5, pady=5)
input_file_label = tk.Label(file_frame, text=": 없음")
input_file_label.grid(row=0, column=1, sticky='w')

output_file_button = tk.Button(file_frame, text="출력 파일 선택", command=select_output_file)
output_file_button.grid(row=1, column=0, padx=5, pady=5)
output_file_label = tk.Label(file_frame, text=": 없음")
output_file_label.grid(row=1, column=1, sticky='w')

convert_button = tk.Button(main_frame, text="내역서 생성", command=process_files,
                          font=("Helvetica", 12, "bold"), bg="lightblue",
                          width=20, height=2)
convert_button.pack(pady=10)

root.mainloop()
